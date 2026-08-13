"""Idempotently import collaborating workshops from lager.xlsx."""

import argparse
from datetime import UTC, date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.importers.catalogue import boolean_value, clean, normalize
from app.models import Workshop

SHEET_NAME = "Våra Verkstäder"


def text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean(str(value))


def boolean(value: object) -> bool | None:
    if isinstance(value, bool):
        return value
    return boolean_value(text(value))


def timestamp(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=UTC) if value.tzinfo is None else value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=UTC)
    return None


def import_workshops(path: Path, session: Session, sheet_name: str = SHEET_NAME) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found")
    imported = 0
    for row_number, row in enumerate(
        workbook[sheet_name].iter_rows(min_row=2, values_only=True), start=2
    ):
        row = (*row, *(None for _ in range(max(0, 21 - len(row)))))
        name = text(row[1])
        if not name:
            continue
        workshop = session.scalar(
            select(Workshop).where(
                Workshop.source == "lager_xlsx",
                Workshop.source_sheet == sheet_name,
                Workshop.source_row == row_number,
            )
        )
        if workshop is None:
            workshop = Workshop(
                name=name,
                normalized_name=normalize(name),
                source="lager_xlsx",
                source_sheet=sheet_name,
                source_row=row_number,
            )
            session.add(workshop)
        workshop.name = name
        workshop.normalized_name = normalize(name)
        workshop.booking_instructions = text(row[0])
        workshop.loan_car_available = boolean(row[2])
        workshop.agreement_terms = text(row[3])
        workshop.workshop_info = text(row[4])
        workshop.contact_person = text(row[5])
        workshop.address = text(row[6])
        workshop.postal_code = text(row[7])
        workshop.city = text(row[8])
        workshop.phone = text(row[9])
        workshop.email = text(row[10])
        workshop.organization_number = text(row[12])
        workshop.discount_terms = text(row[13])
        workshop.internal_owner = text(row[14])
        workshop.written_agreement = boolean(row[15])
        workshop.terms_updated_at = timestamp(row[16])
        workshop.current_info = text(row[17])
        workshop.is_active = boolean(row[18]) is True
        workshop.restrictions = text(row[19])
        workshop.supports_motorhomes = boolean(row[20])
        imported += 1
    session.commit()
    workbook.close()
    return imported


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet", default=SHEET_NAME)
    args = parser.parse_args()
    with SessionLocal() as session:
        count = import_workshops(args.workbook, session, args.sheet)
    print(f"Workshop import complete: {count} workshops")


if __name__ == "__main__":
    main()
