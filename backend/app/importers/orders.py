"""Idempotently import the order-flow workbook without guessing product links."""

import argparse
import re
import tempfile
import urllib.request
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.importers.catalogue import clean, normalize
from app.models import Customer, Order, OrderItem, ProductIdentifier

SHEET_URL = "https://docs.google.com/spreadsheets/d/1g3neK02qON0TecH9aAt0xxjDfw8uNWmk87XhvckR8pA/export?format=xlsx"
DEFAULT_SHEET = "2024"
MAX_ORDER_LINE_AMOUNT = Decimal("100000000")


def text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean(str(value))


def as_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=UTC) if value.tzinfo is None else value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=UTC)
    return None


def amount(value: object) -> Decimal | None:
    raw = text(value)
    if raw is None:
        return None
    # Price cells contain values such as "1 595 kr + frakt", but occasionally
    # contain notes or phone numbers. Parse one leading amount conservatively and
    # reject implausible values instead of failing the complete import.
    match = re.match(r"^\s*(-?\d[\d .]*(?:,\d{1,2})?)", raw)
    if not match:
        return None
    candidate = match.group(1).replace(" ", "")
    if "," in candidate:
        candidate = candidate.replace(".", "").replace(",", ".")
    elif candidate.count(".") > 1 or (
        candidate.count(".") == 1 and len(candidate.rsplit(".", 1)[1]) == 3
    ):
        candidate = candidate.replace(".", "")
    try:
        parsed = Decimal(candidate)
    except InvalidOperation:
        return None
    return parsed if Decimal(0) <= parsed < MAX_ORDER_LINE_AMOUNT else None


def canonical_status(workflow: str | None, sheet_status: str | None) -> str:
    combined = f"{workflow or ''} {sheet_status or ''}".casefold()
    if "avbok" in combined:
        return "cancelled"
    if "strul" in combined or "reklamation" in combined:
        return "attention"
    if "klar" in combined or sheet_status == "OK":
        return "completed"
    if "rest" in combined:
        return "backorder"
    return "in_progress"


def resolve_product(session: Session, sku: str | None):
    normalized = normalize(sku)
    if not normalized:
        return None
    identifier = session.scalar(
        select(ProductIdentifier).where(
            ProductIdentifier.kind == "article_number",
            ProductIdentifier.normalized_value == normalized,
        )
    )
    return identifier.product if identifier else None


def import_workbook(
    path: Path, session: Session, sheet_name: str = DEFAULT_SHEET
) -> tuple[int, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet {sheet_name!r} not found; available: {', '.join(workbook.sheetnames)}"
        )
    sheet = workbook[sheet_name]
    imported = unmatched = 0
    seen_orders: set[str] = set()
    for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        external_id = text(row[3])
        customer_name = text(row[4])
        if not external_id or not customer_name:
            continue
        order = session.scalar(
            select(Order).where(Order.source == "google_sheet", Order.external_id == external_id)
        )
        new_order = order is None
        if new_order:
            customer = Customer(name=customer_name)
            session.add(customer)
            session.flush()
            order = Order(
                source="google_sheet",
                external_id=external_id,
                customer_id=customer.id,
                status="in_progress",
                source_sheet=sheet_name,
                source_row=row_number,
                customer=customer,
            )
            session.add(order)
            session.flush()
        customer = order.customer
        customer.name = customer_name
        customer.delivery_address = text(row[5])
        customer.postal_code = text(row[6])
        customer.city = text(row[7])
        customer.phone = text(row[8])
        customer.email = text(row[9])
        workflow_status, sheet_status = text(row[23]), text(row[30]) if len(row) > 30 else None
        order.ordered_at = as_datetime(row[2])
        order.status = canonical_status(workflow_status, sheet_status)
        order.workflow_status = workflow_status
        order.sales_person = text(row[1])
        order.sales_channel = text(row[49]) if len(row) > 49 else None
        row_amount = amount(row[21])
        order.total_amount = (
            (order.total_amount or Decimal(0)) + (row_amount or Decimal(0))
            if external_id in seen_orders
            else row_amount
        )
        order.payment_method = text(row[29]) if len(row) > 29 else None
        order.vehicle_label = text(row[18])
        order.vehicle_year = text(row[19])
        order.registration_number = text(row[20])
        order.shipping_date = as_datetime(row[27])
        order.tracking_number = text(row[24])
        order.notes = text(row[31]) if len(row) > 31 else None
        order.source_row = min(order.source_row, row_number)
        if external_id not in seen_orders:
            order.items.clear()
            if not new_order:
                session.flush()
            seen_orders.add(external_id)
        for kind, description_index, sku_index in (
            ("primary", 12, 14),
            ("electrical", 13, 15),
        ):
            description, sku = text(row[description_index]), text(row[sku_index])
            if not description or description.strip("-") == "":
                continue
            product = resolve_product(session, sku)
            link_status = "linked" if product else ("missing_sku" if not sku else "unmatched")
            unmatched += int(product is None)
            order.items.append(
                OrderItem(
                    position=len(order.items) + 1,
                    kind=kind,
                    source_sku=sku,
                    description=description,
                    product=product,
                    link_status=link_status,
                )
            )
        imported = len(seen_orders)
        if imported % 500 == 0:
            session.flush()
    session.commit()
    workbook.close()
    return imported, unmatched


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", nargs="?", type=Path)
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Workbook tab to import")
    parser.add_argument("--download", action="store_true", help="Download the configured workbook")
    args = parser.parse_args()
    if args.download:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "orders.xlsx"
            urllib.request.urlretrieve(SHEET_URL, path)  # noqa: S310
            run_import(path, args.sheet)
    elif args.workbook:
        run_import(args.workbook, args.sheet)
    else:
        parser.error("provide a workbook or use --download")


def run_import(path: Path, sheet_name: str) -> None:
    with SessionLocal() as session:
        imported, unmatched = import_workbook(path, session, sheet_name)
    print(f"Order import complete: {imported} orders, {unmatched} unmatched items")


if __name__ == "__main__":
    main()
